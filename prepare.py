#!/usr/bin/env python3
"""Prepare the GL261 mouse brain tumor ultrasound dataset (V2.1 split).

Dataset structure (38 recordings, 1856 images, 12 mice):
  data/raw/
    rec__YYYYMMDD_*/
      Images/    (RGB uint8 PNGs, variable sizes ~672x782-796x906)
      Masks/     (RGB uint8 PNGs, {img_name}_mask.png naming)
      ReadMe.xlsx (per-recording metadata: tumor/non-tumor, in/ex-vivo)

Pipeline:
  1. Parse per-recording metadata from ReadMe.xlsx files
  2. Identify mice from recording dates + metadata
  3. Create mouse-level train/val split (9 train / 3 val)
  4. Copy images + binarize masks
  5. Write CSV manifests
  6. Format for nnU-Net (Dataset501_GL261):
     imagesTr/labelsTr = train, imagesTs/labelsTs = val

Usage:
  python prepare.py
  python prepare.py --raw-root data/raw --out-root data/processed
"""

import argparse
import json
import shutil
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from PIL import Image
from tqdm import tqdm

ROOT = Path(__file__).resolve().parent


# ---------------------------------------------------------------------------
# Metadata parsing
# ---------------------------------------------------------------------------

def _parse_recording_metadata(rec_dir: Path) -> dict:
    """Read ReadMe.xlsx from a recording directory."""
    xlsx = rec_dir / "ReadMe.xlsx"
    rec_name = rec_dir.name
    meta = {"rec_id": rec_name, "condition": "unknown", "vivo": "unknown",
            "tumor_size": "", "gender": "", "imaging_plane": ""}

    if not xlsx.exists():
        return meta

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) >= 2:
            data_row = rows[1]
            tumor_str = str(data_row[0] or "").strip().lower()
            vivo_str = str(data_row[1] or "").strip().lower()
            meta["tumor_size"] = str(data_row[2] or "").strip()
            meta["gender"] = str(data_row[3] or "").strip()
            if len(data_row) > 4:
                meta["imaging_plane"] = str(data_row[4] or "").strip()

            is_tumor = "tumor" in tumor_str and "non" not in tumor_str
            is_exvivo = "ex" in vivo_str
            if is_tumor and is_exvivo:
                meta["condition"] = "tumor_exvivo"
            elif is_tumor:
                meta["condition"] = "tumor_iv"
            elif "non" in tumor_str:
                meta["condition"] = "nontumor_iv"
            meta["vivo"] = "ex-vivo" if is_exvivo else "in-vivo"
    except Exception as e:
        print(f"[WARN] Failed to parse {xlsx}: {e}")

    return meta


def _find_image_mask_pairs(rec_dir: Path) -> list[dict]:
    """Find matching image-mask pairs in a recording directory."""
    img_dir = rec_dir / "Images"
    mask_dir = rec_dir / "Masks"
    pairs = []

    if not img_dir.exists():
        return pairs

    for img_path in sorted(img_dir.glob("*.png")):
        mask_name = img_path.stem + "_mask.png"
        mask_path = mask_dir / mask_name

        pairs.append({
            "img_path": img_path,
            "mask_path": mask_path if mask_path.exists() else None,
            "img_name": img_path.name,
        })

    return pairs


# ---------------------------------------------------------------------------
# Mouse identification
# ---------------------------------------------------------------------------

def _identify_mouse(rec_metas: list[dict]) -> dict[str, str]:
    """Assign each recording to a mouse ID based on date + metadata.

    Recordings from the same date with matching tumor size are the same mouse.
    Special case: 2024-09-13 has two mice (3.5mm and 4.5mm), and the 4.5mm
    mouse was re-imaged on 2024-09-20.

    Returns dict mapping rec_id -> mouse_id (e.g., "M01"..."M12").
    """
    def _date(rec_id: str) -> str:
        return rec_id.split("__")[1].split("_")[0]

    def _tumor_mm(size_str: str) -> float:
        s = size_str.replace("mm", "").replace("\u2026", "").strip()
        if not s:
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0

    groups: dict[str, list[str]] = {}

    for m in rec_metas:
        date = _date(m["rec_id"])
        tmm = _tumor_mm(m["tumor_size"])

        # Special case: 4.5mm tumor on 2024-09-13 or 2024-09-20 is same mouse
        if tmm >= 4.0 and date in ("20240913", "20240920"):
            key = "tumor_4.5mm_sept"
        else:
            key = f"{date}_{tmm}"

        groups.setdefault(key, []).append(m["rec_id"])

    sorted_groups = sorted(groups.items(), key=lambda x: min(x[1]))

    rec_to_mouse = {}
    for i, (_, recs) in enumerate(sorted_groups, 1):
        mid = f"M{i:02d}"
        for rid in recs:
            rec_to_mouse[rid] = mid

    return rec_to_mouse


# ---------------------------------------------------------------------------
# Splitting
# ---------------------------------------------------------------------------

VAL_MICE = {"M04", "M09", "M10"}


def _assign_splits(rec_metas: list[dict], rec_to_mouse: dict[str, str]) -> dict[str, str]:
    """Assign each recording to train/val via mouse-level splitting.

    Val mice (deterministic, chosen for condition coverage):
      - M04: tumor_iv (representative)
      - M09: nontumor_iv (measures false positive rate)
      - M10: tumor_exvivo (measures iv vs ev gap)
    """
    rec_splits = {}
    for m in rec_metas:
        mid = rec_to_mouse[m["rec_id"]]
        rec_splits[m["rec_id"]] = "val" if mid in VAL_MICE else "train"
    return rec_splits


# ---------------------------------------------------------------------------
# Image processing
# ---------------------------------------------------------------------------

def _binarize_mask(mask_path: Path) -> np.ndarray:
    """Load RGB mask and binarize to single-channel 0/1."""
    arr = np.array(Image.open(mask_path))
    if arr.ndim == 3:
        binary = (arr.max(axis=2) > 0).astype(np.uint8)
    else:
        binary = (arr > 0).astype(np.uint8)
    return binary


# ---------------------------------------------------------------------------
# nnU-Net formatting
# ---------------------------------------------------------------------------

def _format_nnunet(records: list[dict], out_root: Path) -> dict:
    """Create nnU-Net Dataset501_GL261 directory structure."""
    ds_dir = out_root / "Dataset501_GL261"
    images_tr = ds_dir / "imagesTr"
    labels_tr = ds_dir / "labelsTr"
    images_ts = ds_dir / "imagesTs"
    labels_ts = ds_dir / "labelsTs"

    for d in [images_tr, labels_tr, images_ts, labels_ts]:
        d.mkdir(parents=True, exist_ok=True)

    case_counter = 1
    n_train = 0
    n_test = 0

    for rec in sorted(records, key=lambda r: (r["rec_id"], r["img_name"])):
        case_id = f"GL261_{case_counter:04d}"
        case_counter += 1

        src_img = rec["proc_img_path"]
        src_mask = rec["proc_mask_path"]

        if not src_img.exists():
            continue

        if rec["split"] == "train":
            dst_img = images_tr / f"{case_id}_0000.png"
            dst_mask = labels_tr / f"{case_id}.png"
            n_train += 1
        else:
            dst_img = images_ts / f"{case_id}_0000.png"
            dst_mask = labels_ts / f"{case_id}.png"
            n_test += 1

        img = Image.open(src_img)
        if img.mode != "L":
            img = img.convert("L")
        img.save(dst_img)

        if src_mask.exists():
            binary = _binarize_mask(src_mask)
            Image.fromarray(binary, mode="L").save(dst_mask)
        else:
            empty = np.zeros((img.height, img.width), dtype=np.uint8)
            Image.fromarray(empty, mode="L").save(dst_mask)

    dataset_json = {
        "channel_names": {"0": "ultrasound"},
        "labels": {"background": 0, "tumor": 1},
        "numTraining": n_train,
        "file_ending": ".png",
    }

    with (ds_dir / "dataset.json").open("w") as f:
        json.dump(dataset_json, f, indent=2)

    return {"nnunet_dir": str(ds_dir), "numTraining": n_train, "numTest": n_test}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Prepare GL261 dataset")
    parser.add_argument("--raw-root", type=Path, default=ROOT / "data" / "raw")
    parser.add_argument("--out-root", type=Path, default=ROOT / "data" / "processed")
    parser.add_argument("--skip-nnunet", action="store_true",
                        help="Skip nnU-Net directory formatting")
    args = parser.parse_args()

    raw_root = args.raw_root.resolve()
    out_root = args.out_root.resolve()

    if not raw_root.exists():
        print(f"[ERROR] Raw root does not exist: {raw_root}", file=sys.stderr)
        sys.exit(1)

    # --- Step 1: Parse all recordings ---
    print("--- Step 1: Parse recording metadata ---")
    rec_dirs = sorted([d for d in raw_root.iterdir()
                       if d.is_dir() and d.name.startswith("rec__")])
    print(f"Found {len(rec_dirs)} recordings")

    rec_metas = []
    all_records = []

    for rec_dir in rec_dirs:
        meta = _parse_recording_metadata(rec_dir)
        pairs = _find_image_mask_pairs(rec_dir)
        rec_metas.append(meta)

        for pair in pairs:
            has_tumor = False
            tumor_area_pct = 0.0
            if pair["mask_path"] is not None:
                try:
                    mask_arr = _binarize_mask(pair["mask_path"])
                    has_tumor = bool(np.any(mask_arr > 0))
                    tumor_area_pct = float(np.count_nonzero(mask_arr) / mask_arr.size * 100)
                except Exception:
                    pass

            all_records.append({
                "rec_id": meta["rec_id"],
                "condition": meta["condition"],
                "vivo": meta["vivo"],
                "tumor_size": meta["tumor_size"],
                "gender": meta["gender"],
                "img_name": pair["img_name"],
                "raw_img_path": pair["img_path"],
                "raw_mask_path": pair["mask_path"],
                "has_tumor": int(has_tumor),
                "tumor_area_pct": round(tumor_area_pct, 4),
            })

        print(f"  {meta['rec_id']}: {meta['condition']:15s} {meta['vivo']:8s} "
              f"tumor_size={meta['tumor_size']:6s} {len(pairs):3d} images")

    print(f"\nTotal images: {len(all_records)}")

    # --- Step 2: Identify mice and assign splits ---
    print("\n--- Step 2: Identify mice ---")
    rec_to_mouse = _identify_mouse(rec_metas)
    n_mice = len(set(rec_to_mouse.values()))
    print(f"Identified {n_mice} mice from {len(rec_metas)} recordings")

    mouse_info: dict[str, dict] = {}
    for m in rec_metas:
        mid = rec_to_mouse[m["rec_id"]]
        if mid not in mouse_info:
            mouse_info[mid] = {"conditions": set(), "recs": [], "tumor_size": m["tumor_size"],
                               "gender": m["gender"]}
        mouse_info[mid]["conditions"].add(m["condition"])
        mouse_info[mid]["recs"].append(m["rec_id"])

    for mid in sorted(mouse_info):
        info = mouse_info[mid]
        n_imgs = sum(1 for r in all_records if rec_to_mouse[r["rec_id"]] == mid)
        conds = "+".join(sorted(info["conditions"]))
        print(f"  {mid}: {len(info['recs']):2d} recs, {n_imgs:4d} imgs, "
              f"cond={conds}, size={info['tumor_size']}, sex={info['gender']}")

    print("\n--- Step 3: Assign mouse-level splits ---")
    split_assignments = _assign_splits(rec_metas, rec_to_mouse)

    for rec in all_records:
        rec["mouse_id"] = rec_to_mouse[rec["rec_id"]]
        rec["split"] = split_assignments[rec["rec_id"]]

    df = pd.DataFrame(all_records)
    for split in ["train", "val"]:
        sdf = df[df["split"] == split]
        n_mice_s = sdf["mouse_id"].nunique()
        n_recs = sdf["rec_id"].nunique()
        n_tumor = int(sdf["has_tumor"].sum())
        conds = ", ".join(sorted(sdf["condition"].unique()))
        print(f"  {split:6s}: {n_mice_s} mice, {n_recs:2d} recs, {len(sdf):4d} images "
              f"({n_tumor} tumor+, {len(sdf) - n_tumor} tumor-) [{conds}]")

    # Leakage check
    mice_train = set(df[df["split"] == "train"]["mouse_id"])
    mice_val = set(df[df["split"] == "val"]["mouse_id"])
    overlap = mice_train & mice_val
    if overlap:
        print(f"  [ERROR] Mouse leakage between train and val: {overlap}")
        sys.exit(1)
    print("  Leakage check: PASSED (no mouse in multiple splits)")

    # --- Step 4: Copy images and binarize masks ---
    print("\n--- Step 4: Copy images and process masks ---")
    png_out = out_root / "png"
    mask_out = out_root / "masks"
    png_out.mkdir(parents=True, exist_ok=True)
    mask_out.mkdir(parents=True, exist_ok=True)

    for rec in tqdm(all_records, desc="Processing"):
        out_fname = f"{rec['rec_id']}__{rec['img_name']}"
        dst_img = png_out / out_fname
        dst_mask = mask_out / out_fname

        rec["proc_img_path"] = dst_img
        rec["proc_mask_path"] = dst_mask
        rec["proc_fname"] = out_fname

        if not dst_img.exists():
            shutil.copy2(rec["raw_img_path"], dst_img)

        if rec["raw_mask_path"] is not None and not dst_mask.exists():
            binary = _binarize_mask(rec["raw_mask_path"])
            Image.fromarray(binary * 255, mode="L").save(dst_mask)
        elif not dst_mask.exists():
            img = Image.open(rec["raw_img_path"])
            empty = np.zeros((img.height, img.width), dtype=np.uint8)
            Image.fromarray(empty, mode="L").save(dst_mask)

    print(f"Processed {len(all_records)} images to {png_out}")

    # --- Step 5: Write CSV manifests ---
    print("\n--- Step 5: Write CSV manifests ---")
    csv_out = out_root / "csv"
    csv_out.mkdir(parents=True, exist_ok=True)

    manifest_df = pd.DataFrame([{
        "filename": r["proc_fname"],
        "rec_id": r["rec_id"],
        "mouse_id": r["mouse_id"],
        "condition": r["condition"],
        "split": r["split"],
        "has_tumor": r["has_tumor"],
        "tumor_area_pct": r["tumor_area_pct"],
    } for r in all_records])
    manifest_df.to_csv(csv_out / "manifest_full.csv", index=False)
    print(f"  manifest_full.csv: {len(manifest_df)} rows")

    inv_rows = []
    for m in rec_metas:
        mid = rec_to_mouse[m["rec_id"]]
        recs = [r for r in all_records if r["rec_id"] == m["rec_id"]]
        n_tumor = sum(r["has_tumor"] for r in recs)
        inv_rows.append({
            "rec_id": m["rec_id"],
            "mouse_id": mid,
            "condition": m["condition"],
            "vivo": m["vivo"],
            "tumor_size": m["tumor_size"],
            "gender": m["gender"],
            "split": split_assignments[m["rec_id"]],
            "n_images": len(recs),
            "n_tumor_positive": n_tumor,
            "n_tumor_negative": len(recs) - n_tumor,
        })
    inv_df = pd.DataFrame(inv_rows)
    inv_df.to_csv(csv_out / "recording_inventory.csv", index=False)
    print(f"  recording_inventory.csv: {len(inv_df)} rows")

    # --- Step 6: nnU-Net formatting ---
    if not args.skip_nnunet:
        print("\n--- Step 6: Format for nnU-Net (Dataset501_GL261) ---")
        nnunet_root = out_root / "nnunet"
        nnunet_stats = _format_nnunet(all_records, nnunet_root)
        print(f"  nnU-Net dir: {nnunet_stats['nnunet_dir']}")
        print(f"  Training cases: {nnunet_stats['numTraining']}")
        print(f"  Test cases: {nnunet_stats['numTest']}")
    else:
        print("\n--- Step 6: Skipped (--skip-nnunet) ---")
        nnunet_stats = None

    # --- Summary ---
    print("\n" + "=" * 70)
    print("GL261 V2.1 PREPARATION COMPLETE")
    print("=" * 70)

    for split in ["train", "val"]:
        sdf = df[df["split"] == split]
        n_mice_s = sdf["mouse_id"].nunique()
        n_recs = sdf["rec_id"].nunique()
        n_tumor = int(sdf["has_tumor"].sum())
        print(f"  {split:6s}: {n_mice_s} mice, {n_recs:2d} recs, {len(sdf):4d} images "
              f"({n_tumor} tumor+, {len(sdf) - n_tumor} tumor-)")

    if nnunet_stats:
        print(f"\n  nnU-Net: {nnunet_stats['numTraining']} imagesTr, "
              f"{nnunet_stats['numTest']} imagesTs")
    print(f"\n  Output: {out_root}")


if __name__ == "__main__":
    main()
